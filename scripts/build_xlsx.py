#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

LOGGER = logging.getLogger(__name__)

STAGE_SHEETS = (
    ("add_requests", ("add_requests",)),
    ("prepare_inputs", ("prepare_inputs",)),
    (
        "prepare_attn",
        ("prepare_attn_runner", "prepare_attn_model_state"),
    ),
    ("run_fullgraph", ("run_fullgraph",)),
    ("sample", ("sample",)),
    ("output", ("async_output_init", "postprocess_sampled")),
)
BASE_HEADERS = (
    "函数",
    "序号",
    "全局调用",
    "统计范围",
    "时间(us)",
    "time_enabled",
    "time_running",
    "valid",
)
SUMMARY_METRICS = (
    "CPU利用率",
    "time(us)",
    "cycles",
    "cycle占比",
    "instructions",
    "IPC",
    "Retire",
    "FrontendBound",
    "BackendBound",
    "BadSpec",
    "dp_spec",
    "ld_spec",
    "st_spec",
    "branch_spec",
    "ase_spec",
    "br missrate",
    "br mpki",
    "l1i missrate",
    "l1i mpki",
    "l2i missrate",
    "l2i mpki",
    "l1d missrate",
    "l1d mpki",
    "L2d missrate",
    "L2d mpki",
    "L3 missrate",
    "L3 mpki",
    "itlb missrate",
    "itlb mpki",
    "dtlb missrate",
    "dtlb mpki",
    "stlb missrate",
    "stlb mpki",
)
FORMULA_TOKEN = re.compile(r"\{([^{}]+)\}")
SAFE_SEGMENT = re.compile(r"[A-Za-z0-9._-]+")
NUMBER = re.compile(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)")

HEADER_FILL = "FF1F4E78"
HEADER_BORDER_COLOR = "FFB4C6E7"
BODY_BORDER_COLOR = "FFD9E2F3"
SUMMARY_LABEL_FILL = "FFD9EAF7"
SUMMARY_LABEL_FONT = "FF17365D"
WHITE = "FFFFFFFF"
BLACK = "FF000000"

HEADER_BORDER = Border(
    left=Side(style="thin", color=HEADER_BORDER_COLOR),
    right=Side(style="thin", color=HEADER_BORDER_COLOR),
    top=Side(style="thin", color=HEADER_BORDER_COLOR),
    bottom=Side(style="thin", color=HEADER_BORDER_COLOR),
)
BODY_BORDER = Border(
    left=Side(style="thin", color=BODY_BORDER_COLOR),
    right=Side(style="thin", color=BODY_BORDER_COLOR),
    top=Side(style="thin", color=BODY_BORDER_COLOR),
    bottom=Side(style="thin", color=BODY_BORDER_COLOR),
)


@dataclass(frozen=True)
class DerivedMetric:
    name: str
    number_format: str
    formula: str


@dataclass(frozen=True)
class GroupSpec:
    name: str
    suffix: str
    events: tuple[str, ...]
    event_headers: tuple[str, ...]
    semantic_headers: tuple[str, ...]
    derived: tuple[DerivedMetric, ...]


@dataclass(frozen=True)
class StageRow:
    sequence: int
    global_call: int
    wall_time_us: float | None
    time_enabled: int
    time_running: int
    valid: int
    scope: str
    counts: dict[str, int]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate the final vLLM PMU Excel workbook."
    )
    parser.add_argument("run_root", type=Path)
    parser.add_argument("--config", type=Path, required=True)
    parser.add_argument("--chip", required=True)
    parser.add_argument("--version", default="0.26")
    parser.add_argument("--model-short", default="qwen3")
    parser.add_argument("--input-len", type=int, default=7000)
    parser.add_argument("--output-len", type=int, default=100)
    return parser.parse_args()


def require_mapping(value: object, context: str) -> dict[str, object]:
    if not isinstance(value, dict) or not all(isinstance(key, str) for key in value):
        raise ValueError(f"{context} must be an object")
    return value


def require_string(value: object, context: str) -> str:
    if not isinstance(value, str) or not value:
        raise ValueError(f"{context} must be a non-empty string")
    return value


def require_string_list(value: object, context: str) -> tuple[str, ...]:
    if not isinstance(value, list) or not all(
        isinstance(item, str) and item for item in value
    ):
        raise ValueError(f"{context} must be a list of non-empty strings")
    return tuple(value)


def load_config(path: Path) -> tuple[GroupSpec, ...]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    root = require_mapping(payload, str(path))
    groups_value = root.get("groups")
    if not isinstance(groups_value, list) or not groups_value:
        raise ValueError(f"{path}: groups must be a non-empty list")

    groups: list[GroupSpec] = []
    for group_index, group_value in enumerate(groups_value):
        context = f"{path}: groups[{group_index}]"
        group = require_mapping(group_value, context)
        name = require_string(group.get("name"), f"{context}.name")
        suffix = require_string(group.get("suffix"), f"{context}.suffix")
        events = require_string_list(group.get("events"), f"{context}.events")
        event_headers_value = group.get("event_headers", list(events))
        event_headers = require_string_list(
            event_headers_value, f"{context}.event_headers"
        )
        if len(event_headers) != len(events):
            raise ValueError(f"{context}: event_headers length must match events")
        semantic_value = group.get("semantic_headers", [])
        semantic_headers = require_string_list(
            semantic_value, f"{context}.semantic_headers"
        )
        if semantic_headers and len(semantic_headers) != len(events):
            raise ValueError(f"{context}: semantic_headers length must match events")

        derived_value = group.get("derived")
        if not isinstance(derived_value, list):
            raise ValueError(f"{context}.derived must be a list")
        derived: list[DerivedMetric] = []
        available_tokens = set(events)
        for metric_index, metric_value in enumerate(derived_value):
            metric_context = f"{context}.derived[{metric_index}]"
            metric = require_mapping(metric_value, metric_context)
            metric_name = require_string(metric.get("name"), f"{metric_context}.name")
            number_format = require_string(
                metric.get("number_format"),
                f"{metric_context}.number_format",
            )
            formula = require_string(metric.get("formula"), f"{metric_context}.formula")
            unknown_tokens = set(FORMULA_TOKEN.findall(formula)) - available_tokens
            if unknown_tokens:
                raise ValueError(
                    f"{metric_context}: unknown formula tokens {sorted(unknown_tokens)}"
                )
            derived.append(DerivedMetric(metric_name, number_format, formula))
            available_tokens.add(metric_name)
        groups.append(
            GroupSpec(
                name=name,
                suffix=suffix,
                events=events,
                event_headers=event_headers,
                semantic_headers=semantic_headers,
                derived=tuple(derived),
            )
        )

    names = [group.name for group in groups]
    suffixes = [group.suffix for group in groups]
    if len(names) != len(set(names)) or len(suffixes) != len(set(suffixes)):
        raise ValueError(f"{path}: group names and suffixes must be unique")
    return tuple(groups)


def read_dict_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [
            {key: value or "" for key, value in row.items() if key is not None}
            for row in csv.DictReader(handle)
        ]


def read_matrix(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [list(row) for row in csv.reader(handle)]


def load_wall_times(run_root: Path, stage: str) -> dict[int, float]:
    rows = read_dict_rows(run_root / "time" / "raw" / f"{stage}.csv")
    return {
        int(row["sequence"]): float(row["wall_time_us"])
        for row in rows
        if row["valid"] == "1"
    }


def load_stage_rows(run_root: Path, group: GroupSpec, stage: str) -> list[StageRow]:
    raw_path = run_root / group.name / "raw" / f"{stage}.csv"
    parsed_path = run_root / group.name / "parsed" / f"{stage}.csv"
    raw_rows = read_dict_rows(raw_path)
    parsed_rows = read_dict_rows(parsed_path)
    if not parsed_rows:
        raise ValueError(f"{parsed_path}: no aligned decode rows")
    decode_calls = {int(row["global_call"]) for row in parsed_rows}
    first_decode_call = min(decode_calls)
    wall_times = load_wall_times(run_root, stage)

    rows: list[StageRow] = []
    for raw in raw_rows:
        sequence = int(raw["sequence"])
        global_call = int(raw["global_call"])
        if global_call in decode_calls:
            scope = "Decode（计入汇总）"
        elif global_call < first_decode_call:
            scope = "Prefill（不计入汇总）"
        else:
            scope = "非对齐（不计入汇总）"
        counts = {event: int(raw[event]) for event in group.events}
        rows.append(
            StageRow(
                sequence=sequence,
                global_call=global_call,
                wall_time_us=wall_times.get(sequence),
                time_enabled=int(raw["time_enabled"]),
                time_running=int(raw["time_running"]),
                valid=int(raw["valid"]),
                scope=scope,
                counts=counts,
            )
        )
    if not rows:
        raise ValueError(f"{raw_path}: no raw rows")
    return rows


def style_detail_header(cell: Cell) -> None:
    cell.font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.border = HEADER_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_detail_body(cell: Cell) -> None:
    cell.font = Font(name="微软雅黑", size=11, color=BLACK)
    cell.fill = PatternFill("solid", fgColor=WHITE)
    cell.border = BODY_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")


def protect_hex_header(value: str) -> str:
    if re.fullmatch(r"0x[0-9a-f]+", value, flags=re.IGNORECASE):
        return f"\u2060{value}"
    return value


def render_formula(
    template: str,
    row_number: int,
    references: dict[str, str],
) -> str:
    formula = template
    for token in FORMULA_TOKEN.findall(template):
        formula = formula.replace(f"{{{token}}}", f"{references[token]}{row_number}")
    return formula


def write_detail_section(
    worksheet: Worksheet,
    run_root: Path,
    group: GroupSpec,
    stage: str,
    start_row: int,
) -> int:
    rows = load_stage_rows(run_root, group, stage)
    header_rows = 2 if group.semantic_headers else 1
    body_start = start_row + header_rows
    headers = (
        *BASE_HEADERS,
        *(protect_hex_header(value) for value in group.event_headers),
        *(metric.name for metric in group.derived),
    )

    if group.semantic_headers:
        semantic = [None] * len(headers)
        for index, value in enumerate(group.semantic_headers, len(BASE_HEADERS)):
            semantic[index] = value
        for column, value in enumerate(semantic, 1):
            cell = worksheet.cell(start_row, column, value)
            style_detail_header(cell)
        worksheet.row_dimensions[start_row].height = 30
        header_row = start_row + 1
    else:
        header_row = start_row

    for column, value in enumerate(headers, 1):
        cell = worksheet.cell(header_row, column, value)
        style_detail_header(cell)
    worksheet.row_dimensions[header_row].height = 30

    event_start = len(BASE_HEADERS) + 1
    derived_start = event_start + len(group.events)
    event_references = {
        event: get_column_letter(event_start + index)
        for index, event in enumerate(group.events)
    }
    derived_references = {
        metric.name: get_column_letter(derived_start + index)
        for index, metric in enumerate(group.derived)
    }
    references = {**event_references, **derived_references}

    for offset, row in enumerate(rows):
        row_number = body_start + offset
        values: list[object] = [
            stage,
            row.sequence,
            row.global_call,
            row.scope,
            row.wall_time_us,
            row.time_enabled,
            row.time_running,
            row.valid,
            *(row.counts[event] for event in group.events),
        ]
        for column, value in enumerate(values, 1):
            cell = worksheet.cell(row_number, column, value)
            style_detail_body(cell)
        for index, metric in enumerate(group.derived, derived_start):
            cell = worksheet.cell(
                row_number,
                index,
                render_formula(metric.formula, row_number, references),
            )
            style_detail_body(cell)
            cell.number_format = metric.number_format
        worksheet.row_dimensions[row_number].height = 21

        worksheet.cell(row_number, 2).number_format = "0"
        worksheet.cell(row_number, 3).number_format = "0"
        worksheet.cell(row_number, 5).number_format = "0.00"
        for column in range(6, derived_start):
            worksheet.cell(row_number, column).number_format = "0"

    return body_start + len(rows) + 1


def size_detail_sheet(worksheet: Worksheet, last_column: int) -> None:
    widths = {"A": 28, "B": 8, "C": 12, "D": 22, "E": 12, "F": 16, "G": 16, "H": 8}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    for column in range(9, last_column + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 14
    worksheet.freeze_panes = "B2"
    worksheet.sheet_view.showGridLines = False


def normalize_summary_rows(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        raise ValueError("summary.csv is empty")
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    labels = [
        row[0]
        for row in normalized[1:]
        if row[0] and row[0] not in {"热点函数", "热点函数占比："}
    ]
    if labels != list(SUMMARY_METRICS):
        raise ValueError("summary metrics do not match the 920b template")

    hotspot_index = next(
        (
            index
            for index, row in enumerate(normalized)
            if row and row[0] in {"热点函数", "热点函数占比："}
        ),
        None,
    )
    if hotspot_index is not None and hotspot_index > 0:
        previous = normalized[hotspot_index - 1]
        if any(value for value in previous):
            normalized.insert(hotspot_index, [""] * width)
    return normalized


def summary_value(value: str) -> tuple[object, str]:
    stripped = value.strip()
    if not stripped:
        return None, "General"
    if stripped.endswith("%") and NUMBER.fullmatch(stripped[:-1]):
        return float(stripped[:-1]) / 100, "0.00%"
    if NUMBER.fullmatch(stripped):
        return float(stripped), "0.00"
    return stripped, "General"


def write_summary(worksheet: Worksheet, run_root: Path) -> None:
    rows = normalize_summary_rows(read_matrix(run_root / "summary.csv"))
    width = len(rows[0])
    for row_index, row in enumerate(rows, 1):
        for column_index, raw_value in enumerate(row, 1):
            value, number_format = summary_value(raw_value)
            cell = worksheet.cell(row_index, column_index, value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_index == 1:
                cell.font = Font(name="Carlito", size=11, bold=True, color=WHITE)
                cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
                cell.border = HEADER_BORDER
            elif column_index == 1:
                cell.font = Font(
                    name="Carlito",
                    size=11,
                    bold=True,
                    color=SUMMARY_LABEL_FONT,
                )
                cell.fill = PatternFill("solid", fgColor=SUMMARY_LABEL_FILL)
                cell.border = HEADER_BORDER
            else:
                cell.font = Font(name="Carlito", size=11)
                cell.border = BODY_BORDER
                cell.number_format = number_format

    worksheet.row_dimensions[1].height = 24
    max_label = max(len(str(row[0])) for row in rows if row)
    worksheet.column_dimensions["A"].width = max(23, min(36, max_label + 2))
    stage_widths = {
        "prepare_attn_model_state": 24,
        "postprocess_sampled": 20,
    }
    for column_index in range(2, width + 1):
        header = str(rows[0][column_index - 1])
        column = get_column_letter(column_index)
        worksheet.column_dimensions[column].width = stage_widths.get(header, 17)
    worksheet.freeze_panes = "B2"
    worksheet.sheet_view.showGridLines = False


def hotspot_path(run_root: Path) -> Path:
    symbolized = run_root / "hotspot" / "perf_report_container_symbols.txt"
    if symbolized.is_file():
        return symbolized
    return run_root / "hotspot" / "perf_report.txt"


def write_hotspot(worksheet: Worksheet, run_root: Path) -> None:
    lines = hotspot_path(run_root).read_text(errors="replace").splitlines()
    header = worksheet.cell(1, 1, "perf report 容器内符号解析输出")
    header.font = Font(name="Carlito", size=11, bold=True, color=WHITE)
    header.fill = PatternFill("solid", fgColor=HEADER_FILL)
    header.border = HEADER_BORDER
    header.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 24
    for row_index, line in enumerate(lines, 2):
        cell = worksheet.cell(row_index, 1, line)
        cell.font = Font(name="Menlo", size=9, color="FF222222")
        cell.alignment = Alignment(
            horizontal="left", vertical="bottom", wrap_text=False
        )
    worksheet.column_dimensions["A"].width = 120
    worksheet.freeze_panes = "B2"
    worksheet.sheet_view.showGridLines = False


def safe_segment(value: str, name: str) -> str:
    if SAFE_SEGMENT.fullmatch(value) is None:
        raise ValueError(f"{name} contains unsafe filename characters: {value!r}")
    return value


def output_path(args: argparse.Namespace) -> Path:
    if args.input_len <= 0 or args.output_len <= 0:
        raise ValueError("input and output lengths must be positive")
    input_tag = (
        f"{args.input_len // 1000}k"
        if args.input_len % 1000 == 0
        else str(args.input_len)
    )
    chip = safe_segment(args.chip, "chip")
    version = safe_segment(args.version, "version")
    model_short = safe_segment(args.model_short, "model-short")
    return args.run_root / (
        f"{chip}_vllm{version}_{model_short}_{input_tag}{args.output_len}.xlsx"
    )


def expected_sheet_names(groups: tuple[GroupSpec, ...]) -> list[str]:
    return [
        "汇总",
        "热点函数",
        *(f"{stem} {group.suffix}" for group in groups for stem, _ in STAGE_SHEETS),
    ]


def validate_saved_workbook(path: Path, groups: tuple[GroupSpec, ...]) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        expected = expected_sheet_names(groups)
        if workbook.sheetnames != expected:
            raise ValueError(f"unexpected worksheet order in {path}")
        for worksheet in workbook.worksheets:
            if worksheet.freeze_panes != "B2":
                raise ValueError(f"{path}: {worksheet.title} freeze pane is not B2")
        if workbook["热点函数"].max_row < 2:
            raise ValueError(f"{path}: hotspot report is empty")
    finally:
        workbook.close()


def build_workbook(args: argparse.Namespace, groups: tuple[GroupSpec, ...]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("汇总")
    hotspot = workbook.create_sheet("热点函数")
    write_summary(summary, args.run_root)
    write_hotspot(hotspot, args.run_root)

    for group in groups:
        last_column = len(BASE_HEADERS) + len(group.events) + len(group.derived)
        for stem, stages in STAGE_SHEETS:
            worksheet = workbook.create_sheet(f"{stem} {group.suffix}")
            start_row = 1
            for stage in stages:
                start_row = write_detail_section(
                    worksheet,
                    args.run_root,
                    group,
                    stage,
                    start_row,
                )
            size_detail_sheet(worksheet, last_column)

    workbook.active = 0
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = output_path(args)
    workbook.save(output)
    workbook.close()
    validate_saved_workbook(output, groups)
    return output


def main() -> int:
    args = parse_args()
    groups = load_config(args.config)
    output = build_workbook(args, groups)
    LOGGER.info("wrote %s", output)
    return 0


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    raise SystemExit(main())
