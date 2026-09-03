# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: Copyright contributors to the vLLM project

import csv
import json
from argparse import Namespace
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pytest import MonkeyPatch, approx

from scripts import build_xlsx
from scripts.hygon_c86_7490 import summary

REPO_ROOT = Path(__file__).resolve().parents[2]
CONFIG_PATH = REPO_ROOT / "scripts" / "hygon_c86_7490" / "report_config.json"
ENV_PATH = REPO_ROOT / "scripts" / "config.env"
RUN_PATH = REPO_ROOT / "scripts" / "hygon_c86_7490" / "run.sh"

EXPECTED_CORE_GROUPS = {
    "topdown": (
        "0x76,0xc0,0xc1,0x03aa,0x0487",
        "cycles,instructions,retired_uops,dispatched_uops,frontend_stall_any",
    ),
    "branch": (
        "0xc0,0xc2,0xc3",
        "instructions,branches,branch_misses",
    ),
    "spec_ls": (
        "0x03aa,0x0129,0x0229,0x0429,0xc2",
        "dispatched_uops,load_ops,store_ops,load_store_ops,branches",
    ),
    "spec_ase": (
        "0x03aa,0x0f00",
        "dispatched_uops,fpu_spec_uops",
    ),
    "icache": (
        "0xc0,0x80,0x81,0x0764,0x0164",
        ("instructions,l1i_fetch_windows,l1i_miss_windows,l2i_accesses,l2i_misses"),
    ),
    "dcache": (
        "0xc0,0x0729,0xc860,0x0864,0xf064",
        "instructions,ls_ops,l1d_miss_proxy,l2d_misses,l2d_hits",
    ),
    "tlb": (
        "0xc0,0x0729,0xff45,0x0f45,0xf045",
        "instructions,ls_ops,l1_dtlb_misses,stlb_hits,stlb_misses",
    ),
}


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def sample_rows() -> dict[str, list[dict[str, str]]]:
    return {
        "time": [
            {"wall_time_us": "10", "thread_cpu_time_us": "10"},
            {"wall_time_us": "90", "thread_cpu_time_us": "0"},
        ],
        "topdown": [
            {
                "cycles": "100",
                "instructions": "100",
                "retired_uops": "300",
                "dispatched_uops": "360",
                "frontend_stall_any": "10",
            },
            {
                "cycles": "900",
                "instructions": "450",
                "retired_uops": "2400",
                "dispatched_uops": "2700",
                "frontend_stall_any": "90",
            },
        ],
        "branch": [
            {"instructions": "10", "branches": "1", "branch_misses": "1"},
            {"instructions": "990", "branches": "99", "branch_misses": "9"},
        ],
        "spec_ls": [
            {
                "dispatched_uops": "1000",
                "load_ops": "200",
                "store_ops": "100",
                "load_store_ops": "50",
                "branches": "100",
            }
        ],
        "spec_ase": [{"dispatched_uops": "1000", "fpu_spec_uops": "100"}],
        "icache": [
            {
                "instructions": "1000",
                "l1i_fetch_windows": "200",
                "l1i_miss_windows": "20",
                "l2i_accesses": "100",
                "l2i_misses": "5",
            }
        ],
        "dcache": [
            {
                "instructions": "1000",
                "ls_ops": "300",
                "l1d_miss_proxy": "30",
                "l2d_misses": "20",
                "l2d_hits": "80",
            }
        ],
        "tlb": [
            {
                "instructions": "1000",
                "ls_ops": "300",
                "l1_dtlb_misses": "15",
                "stlb_hits": "45",
                "stlb_misses": "5",
            }
        ],
        "l3": [{"l3_accesses": "200", "l3_misses": "20"}],
    }


def test_hygon_uses_locked_zen1_groups_and_formulas() -> None:
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    groups = {group["name"]: group for group in config["groups"]}

    assert list(groups) == [*EXPECTED_CORE_GROUPS, "l3"]
    env_lines = ENV_PATH.read_text(encoding="utf-8").splitlines()
    run_lines = RUN_PATH.read_text(encoding="utf-8").splitlines()
    for name, (events, event_names) in EXPECTED_CORE_GROUPS.items():
        suffix = name.upper()
        assert f"EVENTS_HYGON_{suffix}={events}" in env_lines
        assert f"NAMES_HYGON_{suffix}={event_names}" in env_lines
        assert (
            f"{name}|$EVENTS_HYGON_{suffix}|$NAMES_HYGON_{suffix}" in run_lines
        )
        assert len(events.split(",")) <= 5

    assert groups["dcache"]["events"] == [
        "instructions",
        "ls_ops",
        "l1d_miss_proxy",
        "l2d_misses",
        "l2d_hits",
    ]
    for name, (events, _event_names) in EXPECTED_CORE_GROUPS.items():
        assert groups[name]["event_headers"] == events.split(",")
        assert groups[name]["semantic_headers"] == groups[name]["events"]
    assert groups["l3"]["event_headers"] == [
        "0xff04",
        "0x0106",
    ]
    assert "HYGON_L3_PMU_NAME=amd_l3" in env_lines
    assert "EVENTS_HYGON_L3=0xff04,0x0106" in env_lines
    assert "NAMES_HYGON_L3=l3_accesses,l3_misses" in env_lines
    run_text = RUN_PATH.read_text(encoding="utf-8")
    assert "metric_basis=Hygon Zen1 proxy" in run_text
    assert 'KPERF_PMU_NAME="$HYGON_L3_PMU_NAME"' in run_text
    assert 'l3 "$EVENTS_HYGON_L3"' in run_text
    assert '"$NAMES_HYGON_L3" uncore' in run_text

    topdown_formulas = {
        metric["name"]: metric["formula"] for metric in groups["topdown"]["derived"]
    }
    assert topdown_formulas == {
        "IPC": '=IFERROR({instructions}/{cycles},"")',
        "Retire": '=IFERROR({retired_uops}/(6*{cycles}),"")',
        "FrontendBound": '=IFERROR({frontend_stall_any}/{cycles},"")',
        "BadSpec": (
            '=IFERROR(MAX({dispatched_uops}-{retired_uops},0)/(6*{cycles}),"")'
        ),
        "BackendBound": ('=IFERROR(1-{Retire}-{FrontendBound}-{BadSpec},"")'),
    }


def test_hygon_summary_uses_aggregate_ratios(monkeypatch: MonkeyPatch) -> None:
    rows_by_group = sample_rows()
    monkeypatch.setattr(
        summary,
        "read_rows",
        lambda _root, group, _stage: rows_by_group[group],
    )

    metrics = summary.stage_metrics(Path(), "add_requests")

    assert metrics["CPU利用率"] == approx(0.1)
    assert metrics["time(us)"] == approx(50)
    assert metrics["cycles"] == approx(500)
    assert metrics["频率(MHz)"] == approx(10)
    assert metrics["instructions"] == approx(275)
    assert metrics["IPC"] == approx(0.55)
    assert metrics["Retire"] == approx(0.45)
    assert metrics["FrontendBound"] == approx(0.1)
    assert metrics["BadSpec"] == approx(0.06)
    assert metrics["BackendBound"] == approx(0.39)
    assert metrics["dp_spec"] == approx(0.45)
    assert metrics["ld_spec"] == approx(0.2)
    assert metrics["st_spec"] == approx(0.15)
    assert metrics["branch_spec"] == approx(0.1)
    assert metrics["ase_spec"] == approx(0.1)
    assert metrics["br missrate"] == approx(0.1)
    assert metrics["br mpki"] == approx(10)
    assert metrics["l1i missrate"] == approx(0.1)
    assert metrics["l1i mpki"] == approx(20)
    assert metrics["l2i missrate"] == approx(0.05)
    assert metrics["l2i mpki"] == approx(5)
    assert metrics["l1d missrate"] == approx(0.1)
    assert metrics["l1d mpki"] == approx(30)
    assert metrics["L2d missrate"] == approx(0.2)
    assert metrics["L2d mpki"] == approx(20)
    assert metrics["L3 missrate"] == approx(0.1)
    assert metrics["L3 mpki"] == approx(20 / 550 * 1000)
    assert metrics["itlb missrate"] == "未支持"
    assert metrics["itlb mpki"] == "未支持"
    assert metrics["dtlb missrate"] == approx(0.05)
    assert metrics["dtlb mpki"] == approx(15)
    assert metrics["stlb missrate"] == approx(0.1)
    assert metrics["stlb mpki"] == approx(5)

    expected_labels = [
        label
        for label in build_xlsx.SUMMARY_METRICS
        if label != summary.CYCLES_SHARE_METRIC
    ]
    assert list(metrics) == expected_labels


def test_proxy_sums_above_limit_are_invalid() -> None:
    topdown = [
        {
            "cycles": "100",
            "retired_uops": "600",
            "dispatched_uops": "600",
            "frontend_stall_any": "10",
        }
    ]
    spec_ls = [
        {
            "dispatched_uops": "100",
            "load_ops": "50",
            "store_ops": "30",
            "load_store_ops": "0",
            "branches": "20",
        }
    ]
    spec_ase = [{"dispatched_uops": "100", "fpu_spec_uops": "10"}]

    assert set(summary.topdown_metrics(topdown).values()) == {summary.INVALID}
    assert set(summary.spec_metrics(spec_ls, spec_ase).values()) == {summary.INVALID}


def test_summary_places_cycle_share_below_cycles(
    monkeypatch: MonkeyPatch,
    tmp_path: Path,
) -> None:
    stage_cycles = {
        stage: float(index * 100) for index, stage in enumerate(summary.STAGES, start=1)
    }

    def fake_read_rows(
        _root: Path,
        group: str,
        stage: str,
    ) -> list[dict[str, str]]:
        rows = sample_rows()[group]
        if group != "topdown":
            return rows
        cycles = stage_cycles[stage]
        return [
            {
                "cycles": str(cycles),
                "instructions": str(cycles / 2),
                "retired_uops": str(cycles * 2.7),
                "dispatched_uops": str(cycles * 3),
                "frontend_stall_any": str(cycles * 0.1),
            }
        ]

    monkeypatch.setattr(summary, "parse_args", lambda: Namespace(run_root=tmp_path))
    monkeypatch.setattr(summary, "read_rows", fake_read_rows)
    monkeypatch.setattr(summary, "write_quality", lambda _root: None)

    assert summary.main() == 0

    workbook = Workbook()
    worksheet = workbook.active
    build_xlsx.write_summary(worksheet, tmp_path)
    labels = [worksheet.cell(row, 1).value for row in range(1, worksheet.max_row + 1)]
    metric_labels = [label for label in labels if label in build_xlsx.SUMMARY_METRICS]
    assert metric_labels == list(build_xlsx.SUMMARY_METRICS)
    cycles_row = labels.index("cycles") + 1
    share_row = labels.index("cycle占比") + 1
    assert share_row == cycles_row + 1

    total_cycles = sum(stage_cycles.values())
    for column, stage in enumerate(summary.STAGES, start=2):
        cell = worksheet.cell(share_row, column)
        expected = round(stage_cycles[stage] / total_cycles, 4)
        assert cell.value == approx(expected)
        assert cell.number_format == "0.00%"

    workbook.close()


def test_hygon_workbook_generation(
    monkeypatch: MonkeyPatch,
    tmp_path: Path,
) -> None:
    groups = build_xlsx.load_config(CONFIG_PATH)
    rows_by_group = sample_rows()
    single_rows = {group: rows[:1] for group, rows in rows_by_group.items()}

    for stage_index, stage in enumerate(summary.STAGES, start=1):
        timing = {
            "sequence": 1,
            "global_call": stage_index,
            "wall_time_us": 10,
            "thread_cpu_time_us": 8,
            "cpu_utilization": 0.8,
            "valid": 1,
        }
        write_csv(tmp_path / "time" / "raw" / f"{stage}.csv", [timing])
        write_csv(tmp_path / "time" / "parsed" / f"{stage}.csv", [timing])
        for group in groups:
            counts = single_rows[group.name][0]
            row: dict[str, object] = {
                "sequence": 1,
                "global_call": stage_index,
                "time_enabled": 100,
                "time_running": 100,
                "valid": 1,
                **counts,
            }
            write_csv(
                tmp_path / group.name / "raw" / f"{stage}.csv",
                [row],
            )
            write_csv(
                tmp_path / group.name / "parsed" / f"{stage}.csv",
                [row],
            )

    quality_header = "stage,expected_selected,raw,selected,valid,invalid,status\n"
    for group in summary.GROUPS:
        quality_rows = "".join(f"{stage},1,1,1,1,0,ok\n" for stage in summary.STAGES)
        quality_path = tmp_path / group / "collection_quality.csv"
        quality_path.parent.mkdir(parents=True, exist_ok=True)
        quality_path.write_text(quality_header + quality_rows, encoding="utf-8")
    hotspot = tmp_path / "hotspot" / "perf_report.txt"
    hotspot.parent.mkdir()
    hotspot.write_text("# Samples: 1\n100.00% worker\n", encoding="utf-8")

    monkeypatch.setattr(summary, "parse_args", lambda: Namespace(run_root=tmp_path))
    assert summary.main() == 0

    args = Namespace(
        run_root=tmp_path,
        config=CONFIG_PATH,
        chip="hygon_c86_7490",
        version="0.26",
        model_short="qwen3",
        input_len=7000,
        output_len=100,
    )
    output = build_xlsx.build_workbook(args, groups)
    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook.sheetnames == build_xlsx.expected_sheet_names(groups)
        labels = [
            workbook["汇总"].cell(row, 1).value
            for row in range(1, workbook["汇总"].max_row + 1)
        ]
        assert labels.index("频率(MHz)") == labels.index("CPU利用率") + 1
        assert labels.index("cycle占比") == labels.index("cycles") + 1
        assert labels.index("IPC") + 1 == labels.index("Retire")
        itlb_row = labels.index("itlb missrate") + 1
        assert workbook["汇总"].cell(itlb_row, 2).value == "未支持"
        l3_sheet = workbook["add_requests L3"]
        assert l3_sheet.cell(1, 9).value == "l3_accesses"
        assert l3_sheet.cell(1, 10).value == "l3_misses"
        assert l3_sheet.cell(2, 9).value == "\u20600xff04"
        assert l3_sheet.cell(2, 10).value == "\u20600x0106"
        assert l3_sheet.cell(3, 11).value == '=IFERROR(J3/I3,"")'
    finally:
        workbook.close()
